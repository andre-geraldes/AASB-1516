#!/usr/bin/env python

import openpyxl
import json
import sys
#file = sys.argv[1]

workbook = openpyxl.load_workbook('info.xlsx')
worksheet = workbook['Sheet1']

#Tamanho do excel
rows = worksheet.max_row
columns = worksheet.max_column

data1, data2, data3 = [], [], []
value = ''

for current in range (1, columns):
    data1.append ({'name' : worksheet.cell(row = 1, column = current+1).value})

for current in range (1, rows):
    data2.append ({'name' : worksheet.cell(row = current+1, column = 1).value})

"""
for row in range (1, rows):
    for column in range (1, columns):
        current = worksheet.cell (row = row, column = column).value
        if current != '  ':
            if 'MUT' in current:
                value += '1'
            if 'DOWN' in current:
                value += '2'
            if 'AMP' in current:
                value += '3'
            if 'UP' in current:
                value += '4'
            if 'HOMDEL' in current:
                value += '5'
            if(value != ''):
                data3.append ({"source" : row - 1, "target" : column - 1, "value" : int(value)})
"""

#Preparar ficheiro tsv
mut = []
value = ''
for column in range (2, columns):
    for row in range (2, rows):
        current = worksheet.cell(row = row, column = column).value
        if current != '':
            if 'MUT' in current:
                value += '1'
            if 'DOWN' in current:
                value += '2'
            if 'AMP' in current:
                value += '3'
            if 'UP' in current:
                value += '4'
            if 'HOMDEL' in current:
                value += '5'
        if(value != ''):
            mut.append(str(row-1) + "\t"+str(column-1)+"\t"+value)
        else:
            mut.append(str(row-1)+"\t"+str(column-1)+"\t"+"0")
        value = ''

data_string = json.dumps({'gene' : data1, 'caseID' : data2}, indent = 4)
#data_string = json.dumps({'gene' : data1, 'caseID' : data2, 'links' : data3}, indent = 4)


f = open('cancer.json','w')
f.write(data_string)
f.close()

#Criar ficheiro tsv
f = open('cancer.tsv','w')
f.write("row_idx"+"\t"+"col_idx"+"\t"+"value\n")
for k in mut:
    f.write(k+"\n")
f.close()


# Cells mal formatadas
# 2 Cells com MUT, como funciona?
# Diferentes tipos de MUT (Missense, Inframe, Truncating)
# Pdf com resultados diferentes do Excel
